"""
Backfill histórico UNICA desde los dos Excels descargados de unicadata.com.br.

Archivo 2 (biweekly): contiene 13 hojas TB_01..TB_13 con datos quinzenales 2010→2025.
Tablas usadas:
  TB_02  — moagem (caña molida)  cols: safra | quinzena(date) | region | producao_t
  TB_04  — açúcar                cols: safra | quinzena(date) | region | producao_t
  TB_06  — etanol                cols: safra | quinzena(string) | materia-prima | tipo | region | producao_m3
  TB_09  — ATR + mix             cols: safra | quinzena(date) | region | ATR_total | kg_ATR/ton | Mix_Açúcar | Mix_Etanol | ...
  TB_10  — vendas etanol         cols: safra | tipo_etanol | quinzena(date) | total | externo | interno | outros

Región mapping: "SÃO PAULO"→"SP", "DEMAIS ESTADOS"→"OTHER".
CS (Centro-Sul) = SP + OTHER, calculado como suma.
Safra normalización: "10-11" → "2010-2011".

Uso:
    python -m ingestion.unica_import          # dry-run (imprime stats)
    python -m ingestion.unica_import --commit  # graba en DB
"""
import logging
import os
import re
import sys
from datetime import date
from typing import Optional

logger = logging.getLogger(__name__)

# Ruta base OneDrive
_ONEDRIVE = r"C:\Users\alejandro.fernandez\OneDrive - Sugar Global Trading\sgt_trading\data\Brazil Data"
_EXCEL_BIWEEKLY = os.path.join(_ONEDRIVE, "aecb6699e327f58e01262579b9400080.xlsx")

_MONTHS_PT_ABBR = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

_REGION_MAP = {
    "são paulo": "SP",
    "sao paulo": "SP",
    "sp":        "SP",
    "demais estados": "OTHER",
    "outros estados": "OTHER",
    "other":     "OTHER",
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _norm_safra(raw: str) -> Optional[str]:
    """Normaliza safra a formato "YYYY-YYYY". Acepta "10-11", "2010-2011", etc."""
    if not raw:
        return None
    s = str(raw).strip()
    # Formato completo "2010-2011" o "2010/2011"
    m = re.match(r"^(\d{4})[-/](\d{4})$", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Formato corto "10-11" o "10/11"
    m = re.match(r"^(\d{2})[-/](\d{2})$", s)
    if m:
        y1 = int(m.group(1))
        y2 = int(m.group(2))
        base = 2000 if y1 >= 0 else 1900
        return f"{base+y1}-{base+y2}"
    return None


def _norm_region(raw) -> Optional[str]:
    if raw is None:
        return None
    key = str(raw).strip().lower()
    return _REGION_MAP.get(key)


def _parse_quinzena_date(raw) -> Optional[date]:
    """Acepta datetime, date, o string "16/abr" derivando año del safra."""
    if raw is None:
        return None
    if hasattr(raw, "date"):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return None  # strings tipo "16/abr" se resuelven en _parse_tb06 con contexto safra


def _parse_ddmm_date(raw, safra: str) -> Optional[date]:
    """
    Parsea string "DD/MM" (formato numérico, ej: '01/04', '16/03').
    El año se infiere del safra: si mes >= 4 → año_inicio; si mes < 4 → año_fin.
    """
    if raw is None:
        return None
    if hasattr(raw, "date"):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if not m:
        return None
    day = int(m.group(1))
    mon = int(m.group(2))
    try:
        y_start, y_end = (int(p) for p in safra.split("-"))
    except Exception:
        return None
    year = y_start if mon >= 4 else y_end
    try:
        return date(year, mon, day)
    except ValueError:
        return None


def _parse_tb06_date(raw, safra: str) -> Optional[date]:
    """
    TB_06 tiene fechas como string "16/abr", "01/mai".
    El año se infiere del safra: si mes >= 4 → año_inicio; si mes < 4 → año_fin.
    """
    if raw is None:
        return None
    if hasattr(raw, "date"):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip().lower()
    m = re.match(r"(\d{1,2})/([a-z]+)", s)
    if not m:
        return None
    day = int(m.group(1))
    mon = _MONTHS_PT_ABBR.get(m.group(2)[:3])
    if not mon:
        return None
    try:
        y_start, y_end = (int(p) for p in safra.split("-"))
    except Exception:
        return None
    year = y_start if mon >= 4 else y_end
    try:
        return date(year, mon, day)
    except ValueError:
        return None


def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_pct(v) -> Optional[float]:
    """
    Convierte valor de mix% a float en rango 0-100.
    Acepta: '35,77%' → 35.77 | '1.234,56%' → 1234.56 | 0.357 → 35.7 | 35.77 → 35.77
    """
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith("%"):
        try:
            num = s[:-1].strip()
            if "," in num and "." in num:
                # PT formato con miles: "1.234,56" → "1234.56"
                num = num.replace(".", "").replace(",", ".")
            elif "," in num:
                # PT decimal solo: "35,77" → "35.77"
                num = num.replace(",", ".")
            return round(float(num), 4)
        except Exception:
            pass
    try:
        f = float(s.replace(",", "."))
        if 0 < f <= 1.0:
            return round(f * 100, 4)
        return round(f, 4)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Excel readers — una función por tabla
# ---------------------------------------------------------------------------

def _read_tb_with_date(ws) -> dict:
    """
    Genérico para TB_02 y TB_04: cols = safra | quinzena | region | valor.
    Maneja fechas tanto datetime como string 'DD/MM' o 'DD/abr'.
    """
    out = {}
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(v is not None for v in r)]
    if len(non_empty) < 2:
        return out
    for row in non_empty[1:]:
        if len(row) < 4:
            continue
        safra = _norm_safra(row[0])
        if not safra:
            continue
        # Intentar datetime primero; si falla, intentar string DD/MM numérico
        qdate = _parse_quinzena_date(row[1])
        if qdate is None:
            qdate = _parse_ddmm_date(row[1], safra)
        region = _norm_region(row[2])
        val = _to_int(row[3])
        if qdate and region and val is not None:
            out[(safra, qdate, region)] = val
    return out


def _read_tb02(ws) -> dict:
    """TB_02: moagem → {(safra, quinzena_date, region): cane_t}"""
    return _read_tb_with_date(ws)


def _read_tb04(ws) -> dict:
    """TB_04: açúcar → {(safra, quinzena_date, region): sugar_t}"""
    return _read_tb_with_date(ws)


def _read_tb06(ws) -> dict:
    """
    TB_06: etanol → {(safra, quinzena_date, region, tipo): m3}
    tipo: 'anidro'|'hidratado'|'total' (derivado de 'Tipo' column)
    """
    out = {}
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(v is not None for v in r)]
    if len(non_empty) < 2:
        return out
    # cols: safra | quinzena(str) | materia-prima | tipo | region | producao
    for row in non_empty[1:]:
        if len(row) < 6:
            continue
        safra = _norm_safra(row[0])
        if not safra:
            continue
        qdate = _parse_tb06_date(row[1], safra)
        region = _norm_region(row[4])
        val = _to_int(row[5])
        tipo_raw = str(row[3]).strip().lower() if row[3] else ""
        if "anidro" in tipo_raw:
            tipo = "anidro"
        elif "hidratado" in tipo_raw:
            tipo = "hidratado"
        elif "total" in tipo_raw:
            tipo = "total"
        else:
            continue
        if safra and qdate and region and val is not None:
            out[(safra, qdate, region, tipo)] = val
    return out


def _read_tb09(ws) -> dict:
    """
    TB_09: ATR + mix → {(safra, quinzena_date, region): dict}
    Cols: safra|quinzena|region|ATR_total|kg_ATR/ton|Mix_Açúcar(%)|Mix_Etanol(%)|
          Kg_Açúcar/ton|L_etanol/ton|L_anidro/ton|L_hidratado/ton
    """
    out = {}
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(v is not None for v in r)]
    if len(non_empty) < 2:
        return out
    for row in non_empty[1:]:
        if len(row) < 11:
            continue
        safra = _norm_safra(row[0])
        qdate = _parse_quinzena_date(row[1])
        region = _norm_region(row[2])
        if not (safra and qdate and region):
            continue
        out[(safra, qdate, region)] = {
            "atr_kg_ton":           _to_float(row[4]),
            "sugar_mix_pct":        _to_pct(row[5]),
            "eth_mix_pct":          _to_pct(row[6]),
            "liters_eth_ton":       _to_float(row[8]),
            "liters_anidro_ton":    _to_float(row[9]),
            "liters_hidratado_ton": _to_float(row[10]),
        }
    return out


def _read_tb10(ws) -> dict:
    """
    TB_10: vendas etanol → {(safra, quinzena_date, tipo): dict}
    Cols: safra | tipo_etanol | quinzena | total | externo | interno | outros_fins
    No tiene región — aplica a todo CS.
    """
    out = {}
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(v is not None for v in r)]
    if len(non_empty) < 2:
        return out
    for row in non_empty[1:]:
        if len(row) < 7:
            continue
        safra = _norm_safra(row[0])
        tipo_raw = str(row[1]).strip().lower() if row[1] else ""
        qdate = _parse_quinzena_date(row[2])
        total = _to_int(row[3])
        externo = _to_int(row[4])
        interno = _to_int(row[5])
        if not (safra and qdate):
            continue
        if "anidro" in tipo_raw:
            tipo = "anidro"
        elif "hidratado" in tipo_raw:
            tipo = "hidratado"
        elif "total" in tipo_raw or not tipo_raw:
            tipo = "total"
        else:
            tipo = "total"
        out[(safra, qdate, tipo)] = {
            "total": total,
            "externo": externo,
            "interno": interno,
        }
    return out


# ---------------------------------------------------------------------------
# Join y upsert
# ---------------------------------------------------------------------------

def _build_records(tb02, tb04, tb06, tb09, tb10) -> list[dict]:
    """
    Une todos los datos por (safra, quinzena_date, region).
    Genera registros para SP, OTHER, y CS (suma SP+OTHER).
    """
    # Unión de claves primarias (safra, date, region) de TB02 y TB04
    all_keys = set(tb02.keys()) | set(tb04.keys()) | set(tb09.keys())
    # Filtrar solo SP y OTHER (CS se calcula)
    base_keys = {k for k in all_keys if k[2] in ("SP", "OTHER")}

    records_by_key = {}
    for k in base_keys:
        safra, qdate, region = k
        rec = {
            "safra":                safra,
            "quinzena_date":        qdate,
            "region":               region,
            "cane_crushed_t":       tb02.get(k),
            "sugar_t":              tb04.get(k),
            "ethanol_anidro_m3":    tb06.get((safra, qdate, region, "anidro")),
            "ethanol_hidratado_m3": tb06.get((safra, qdate, region, "hidratado")),
            "ethanol_total_m3":     tb06.get((safra, qdate, region, "total")),
            "eth_sales_total_m3":   None,
            "eth_sales_internal_m3": None,
            "eth_sales_external_m3": None,
            "source": "excel_unica",
        }
        atr_mix = tb09.get(k)
        if atr_mix:
            rec.update(atr_mix)
        else:
            rec.update({
                "atr_kg_ton": None, "sugar_mix_pct": None, "eth_mix_pct": None,
                "liters_eth_ton": None, "liters_anidro_ton": None, "liters_hidratado_ton": None,
            })
        records_by_key[k] = rec

    # TB_10 sales (solo total CS, mapeamos a CS key)
    # Buscar en keys la quincena date de CS, pegar ventas
    tb10_by_date = {}
    for (safra, qdate, tipo), vd in tb10.items():
        if tipo == "total":
            tb10_by_date[(safra, qdate)] = vd

    # Calcular CS = SP + OTHER
    # Recopilar todas las quincenas únicas por safra
    safra_dates = {}
    for safra, qdate, region in base_keys:
        safra_dates.setdefault(safra, set()).add(qdate)

    cs_records = []
    for safra, dates in safra_dates.items():
        for qdate in dates:
            sp_rec = records_by_key.get((safra, qdate, "SP"), {})
            ot_rec = records_by_key.get((safra, qdate, "OTHER"), {})
            if not sp_rec and not ot_rec:
                continue

            def _sum(field):
                a = sp_rec.get(field)
                b = ot_rec.get(field)
                if a is None and b is None:
                    return None
                return (a or 0) + (b or 0)

            def _wavg(field, weight_field):
                """Weighted average for per-ton metrics."""
                va = sp_rec.get(field)
                wb_val = sp_rec.get(weight_field)
                vb = ot_rec.get(field)
                wb2 = ot_rec.get(weight_field)
                if va is None and vb is None:
                    return None
                total_w = (wb_val or 0) + (wb2 or 0)
                if total_w == 0:
                    return None
                num = ((va or 0) * (wb_val or 0)) + ((vb or 0) * (wb2 or 0))
                return round(num / total_w, 4)

            sales = tb10_by_date.get((safra, qdate), {})
            cs_rec = {
                "safra":                  safra,
                "quinzena_date":          qdate,
                "region":                 "CS",
                "cane_crushed_t":         _sum("cane_crushed_t"),
                "sugar_t":                _sum("sugar_t"),
                "ethanol_anidro_m3":      _sum("ethanol_anidro_m3"),
                "ethanol_hidratado_m3":   _sum("ethanol_hidratado_m3"),
                "ethanol_total_m3":       _sum("ethanol_total_m3"),
                "atr_kg_ton":             _wavg("atr_kg_ton", "cane_crushed_t"),
                "sugar_mix_pct":          _wavg("sugar_mix_pct", "cane_crushed_t"),
                "eth_mix_pct":            _wavg("eth_mix_pct", "cane_crushed_t"),
                "liters_eth_ton":         _wavg("liters_eth_ton", "cane_crushed_t"),
                "liters_anidro_ton":      _wavg("liters_anidro_ton", "cane_crushed_t"),
                "liters_hidratado_ton":   _wavg("liters_hidratado_ton", "cane_crushed_t"),
                "eth_sales_total_m3":     sales.get("total"),
                "eth_sales_internal_m3":  sales.get("interno"),
                "eth_sales_external_m3":  sales.get("externo"),
                "source": "excel_unica",
            }
            cs_records.append(cs_rec)

    all_records = list(records_by_key.values()) + cs_records
    return all_records


def upsert_records(session, records: list[dict]) -> int:
    """Upsert registros en unica_biweekly. Retorna número de filas insertadas/actualizadas."""
    from sqlalchemy import text
    from sqlalchemy.exc import IntegrityError
    inserted = 0
    skipped_bad = 0
    for rec in records:
        if rec.get("quinzena_date") is None or rec.get("safra") is None:
            continue
        try:
            existing = session.execute(
                text("SELECT id FROM unica_biweekly WHERE safra=:s AND quinzena_date=:d AND region=:r"),
                {"s": rec["safra"], "d": rec["quinzena_date"], "r": rec["region"]},
            ).fetchone()
            if existing:
                session.execute(
                    text("""
                        UPDATE unica_biweekly SET
                            cane_crushed_t=:cane, sugar_t=:sugar,
                            ethanol_anidro_m3=:eth_a, ethanol_hidratado_m3=:eth_h, ethanol_total_m3=:eth_t,
                            atr_kg_ton=:atr, sugar_mix_pct=:smix, eth_mix_pct=:emix,
                            liters_eth_ton=:let, liters_anidro_ton=:lan, liters_hidratado_ton=:lhid,
                            eth_sales_total_m3=:st, eth_sales_internal_m3=:si, eth_sales_external_m3=:se,
                            source=:src
                        WHERE safra=:safra AND quinzena_date=:qdate AND region=:reg
                    """),
                    {
                        "cane": rec.get("cane_crushed_t"), "sugar": rec.get("sugar_t"),
                        "eth_a": rec.get("ethanol_anidro_m3"), "eth_h": rec.get("ethanol_hidratado_m3"),
                        "eth_t": rec.get("ethanol_total_m3"),
                        "atr": rec.get("atr_kg_ton"), "smix": rec.get("sugar_mix_pct"),
                        "emix": rec.get("eth_mix_pct"), "let": rec.get("liters_eth_ton"),
                        "lan": rec.get("liters_anidro_ton"), "lhid": rec.get("liters_hidratado_ton"),
                        "st": rec.get("eth_sales_total_m3"), "si": rec.get("eth_sales_internal_m3"),
                        "se": rec.get("eth_sales_external_m3"), "src": rec.get("source"),
                        "safra": rec["safra"], "qdate": rec["quinzena_date"], "reg": rec["region"],
                    },
                )
            else:
                session.execute(
                    text("""
                        INSERT INTO unica_biweekly (
                            safra, quinzena_date, region,
                            cane_crushed_t, sugar_t,
                            ethanol_anidro_m3, ethanol_hidratado_m3, ethanol_total_m3,
                            atr_kg_ton, sugar_mix_pct, eth_mix_pct,
                            liters_eth_ton, liters_anidro_ton, liters_hidratado_ton,
                            eth_sales_total_m3, eth_sales_internal_m3, eth_sales_external_m3,
                            source
                        ) VALUES (
                            :safra, :qdate, :reg,
                            :cane, :sugar,
                            :eth_a, :eth_h, :eth_t,
                            :atr, :smix, :emix,
                            :let, :lan, :lhid,
                            :st, :si, :se,
                            :src
                        )
                    """),
                    {
                        "safra": rec["safra"], "qdate": rec["quinzena_date"], "reg": rec["region"],
                        "cane": rec.get("cane_crushed_t"), "sugar": rec.get("sugar_t"),
                        "eth_a": rec.get("ethanol_anidro_m3"), "eth_h": rec.get("ethanol_hidratado_m3"),
                        "eth_t": rec.get("ethanol_total_m3"),
                        "atr": rec.get("atr_kg_ton"), "smix": rec.get("sugar_mix_pct"),
                        "emix": rec.get("eth_mix_pct"), "let": rec.get("liters_eth_ton"),
                        "lan": rec.get("liters_anidro_ton"), "lhid": rec.get("liters_hidratado_ton"),
                        "st": rec.get("eth_sales_total_m3"), "si": rec.get("eth_sales_internal_m3"),
                        "se": rec.get("eth_sales_external_m3"), "src": rec.get("source"),
                    },
                )
                inserted += 1
        except Exception as e:
            try:
                session.rollback()
            except Exception:
                pass
            skipped_bad += 1
            logger.warning(
                "unica_import: fila saltada %s/%s/%s — %s: %s",
                rec.get("safra"), rec.get("quinzena_date"), rec.get("region"),
                type(e).__name__, str(e)[:120],
            )
    session.commit()
    if skipped_bad:
        logger.warning("unica_import: %d filas saltadas por datos inválidos", skipped_bad)
    return inserted


def run_excel_backfill(session, commit: bool = True) -> dict:
    """
    Lee los Excels históricos UNICA y hace upsert en unica_biweekly.
    Retorna stats: {total_records, inserted, skipped}
    """
    import openpyxl
    logger.info("UNICA backfill: leyendo %s", _EXCEL_BIWEEKLY)
    wb = openpyxl.load_workbook(_EXCEL_BIWEEKLY, data_only=True)
    sheets = wb.sheetnames
    logger.info("Hojas disponibles: %s", sheets)

    # Leer tablas relevantes
    tb02 = _read_tb02(wb["TB_02"]) if "TB_02" in sheets else {}
    tb04 = _read_tb04(wb["TB_04"]) if "TB_04" in sheets else {}
    tb06 = _read_tb06(wb["TB_06"]) if "TB_06" in sheets else {}
    tb09 = _read_tb09(wb["TB_09"]) if "TB_09" in sheets else {}
    tb10 = _read_tb10(wb["TB_10"]) if "TB_10" in sheets else {}

    logger.info("TB_02=%d  TB_04=%d  TB_06=%d  TB_09=%d  TB_10=%d",
                len(tb02), len(tb04), len(tb06), len(tb09), len(tb10))

    records = _build_records(tb02, tb04, tb06, tb09, tb10)
    logger.info("Registros construidos: %d (SP+OTHER+CS)", len(records))

    if not commit:
        # dry-run
        safras = sorted({r["safra"] for r in records if r["safra"]})
        regions = {r["region"] for r in records}
        logger.info("DRY-RUN: safras=%s  regiones=%s", safras[:5], regions)
        return {"total_records": len(records), "inserted": 0, "skipped": len(records)}

    inserted = upsert_records(session, records)
    return {"total_records": len(records), "inserted": inserted, "skipped": len(records) - inserted}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s")
    commit = "--commit" in sys.argv

    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
    from database import SessionLocal

    with SessionLocal() as sess:
        stats = run_excel_backfill(sess, commit=commit)

    print(f"\nBackfill UNICA Excels:")
    print(f"  Total registros construidos : {stats['total_records']}")
    print(f"  Insertados (nuevos)          : {stats['inserted']}")
    print(f"  Ya existían (skipped)        : {stats['skipped']}")
    if not commit:
        print("\n  [DRY-RUN] Ejecutar con --commit para grabar en DB")
