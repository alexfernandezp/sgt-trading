"""
European Commission Sugar Market Data — XLSX directo.

Fuente: EC Agricultural Market Observatory (AMO)
  Pagina: https://agriculture.ec.europa.eu/data-and-analysis/markets/overviews/
          market-observatories/sugar_en
  XLSX:   sugar-balance-sheet_en.xlsx  (actualizado mensualmente)

La UE es el segundo mayor productor de azucar de remolacha del mundo (~16 Mt).
Temporada: Oct 1 – Sep 30 (misma que marketing year azucar).
Campaña de procesado: Sep-Ene (beet sugar, Europa del Norte/Centro/Este).

Metodologia:
  1. Scraping de la pagina AMO para encontrar la URL actual del XLSX
     (la URL contiene un UUID que cambia en cada publicacion mensual)
  2. Descarga del XLSX con cache local de 30 dias
  3. Parseo de la hoja de balance (`bilan upd <MesAno>`) — fila produccion
     Balance en miles de toneladas wse (white sugar equivalent)
  4. Override vs USDA solo si divergencia > 0.5 Mt y dato fresco (<60d)

Estructura XLSX confirmada (Apr 2026):
  Hoja "bilan upd Apr26" — balance 2017/18–2025/26 en 1000 t wse
    fila "Production" o equivalente, ultima columna = campaña actual
  Produccion 25/26 confirmada: 16,599 kt = 16.60 Mt

Relevancia:
  - UE produce ~15-17 Mt, USDA V2 code = E4
  - Gap historico USDA vs EC: tipicamente <1 Mt
  - Override solo si gap > 0.5 Mt — en 25/26 USDA≈16.0 Mt vs EC≈16.6 Mt → +0.6 Mt
"""
import io
import logging
import re
import tempfile
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

_TIMEOUT = 30
_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

_AMO_SUGAR_PAGE = (
    "https://agriculture.ec.europa.eu/data-and-analysis/markets/overviews/"
    "market-observatories/sugar_en"
)
_XLSX_FILENAME = "sugar-balance-sheet_en.xlsx"

# Cache local
_CACHE_DIR = Path(__file__).parent.parent / "data" / "ec_cache"
_CACHE_FILE = _CACHE_DIR / _XLSX_FILENAME
_CACHE_MAX_DAYS = 30

# Produccion UE plausible: 12-20 Mt
_EU_PROD_MIN = 12.0
_EU_PROD_MAX = 20.0


# ── Descarga ───────────────────────────────────────────────────────────────────

def _get_xlsx_url_from_page() -> Optional[str]:
    """
    Scraping de la pagina AMO para obtener la URL actual del XLSX.
    La URL contiene un UUID que cambia con cada publicacion.
    """
    try:
        import httpx
        r = httpx.get(_AMO_SUGAR_PAGE, headers=_HEADERS, timeout=_TIMEOUT,
                      follow_redirects=True)
        if r.status_code != 200:
            logger.debug("EC AMO page HTTP %d", r.status_code)
            return None

        # Buscar href que contenga sugar-balance-sheet o el documento UUID EC
        # Patron: /document/download/<UUID>_en?filename=sugar-balance-sheet_en.xlsx
        patterns = [
            r'href="(https://agriculture\.ec\.europa\.eu/document/download/[^"]+sugar-balance-sheet[^"]+)"',
            r'href="(/document/download/[^"]+sugar-balance-sheet[^"]+)"',
            r'"(https://[^"]+agriculture\.ec\.europa\.eu[^"]+sugar-balance-sheet[^"]+\.xlsx)"',
        ]
        for pat in patterns:
            m = re.search(pat, r.text, re.IGNORECASE)
            if m:
                url = m.group(1)
                if url.startswith("/"):
                    url = "https://agriculture.ec.europa.eu" + url
                logger.debug("EC XLSX URL encontrada: %s", url)
                return url

        logger.debug("EC AMO: URL del XLSX no encontrada en la pagina")
    except Exception as e:
        logger.debug("EC AMO page fetch: %s", e)
    return None


def _download_xlsx(url: str) -> Optional[bytes]:
    """Descarga el XLSX de la CE y lo retorna como bytes."""
    try:
        import httpx
        headers = {**_HEADERS, "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*"}
        r = httpx.get(url, headers=headers, timeout=_TIMEOUT, follow_redirects=True)
        if r.status_code == 200 and len(r.content) > 10_000:
            return r.content
        logger.debug("EC XLSX download HTTP %d len=%d", r.status_code, len(r.content))
    except Exception as e:
        logger.debug("EC XLSX download: %s", e)
    return None


def _load_cached_xlsx() -> Optional[bytes]:
    """Lee el XLSX del cache local si no es demasiado viejo."""
    if not _CACHE_FILE.exists():
        return None
    age = date.today() - date.fromtimestamp(_CACHE_FILE.stat().st_mtime)
    if age.days > _CACHE_MAX_DAYS:
        logger.debug("EC cache: XLSX de %d dias — demasiado viejo", age.days)
        return None
    try:
        return _CACHE_FILE.read_bytes()
    except Exception:
        return None


def _save_cache(content: bytes) -> None:
    try:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        _CACHE_FILE.write_bytes(content)
    except Exception as e:
        logger.debug("EC cache write: %s", e)


def _get_xlsx_bytes(force_download: bool = False) -> Optional[bytes]:
    """
    Obtiene el XLSX (cache o descarga).
    Orden: cache local → scraping URL → fallback URL fija
    """
    if not force_download:
        cached = _load_cached_xlsx()
        if cached:
            logger.debug("EC XLSX: usando cache local")
            return cached

    # Buscar URL actual en la pagina AMO
    url = _get_xlsx_url_from_page()
    if url:
        content = _download_xlsx(url)
        if content:
            _save_cache(content)
            logger.info("EC XLSX: descargado de %s (%d bytes)", url, len(content))
            return content

    logger.warning("EC Sugar: no se pudo obtener XLSX de la EC")
    return None


# ── Parsing XLSX ───────────────────────────────────────────────────────────────

def _campaign_str(yr: int) -> str:
    """Ej: 2025 → '25/26' o '25-26'"""
    return f"{str(yr)[2:]}/{str(yr+1)[2:]}"


def _parse_balance_sheet(xlsx_bytes: bytes, campaign_year: int) -> Optional[float]:
    """
    Parsea el XLSX de la CE para extraer produccion EU de la campaña indicada.

    Busca en hojas cuyo nombre contiene "bilan" (balance sheet) o "product".
    La fila de produccion esta etiquetada "Production" o similar.
    Valores en miles de toneladas wse → convertir a Mt.

    Retorna produccion en Mt o None.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("EC Sugar: openpyxl no instalado — pip install openpyxl")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.debug("EC XLSX openpyxl open: %s", e)
        return None

    yr_short = str(campaign_year)[2:]          # "25" para 2025
    yr_next  = str(campaign_year + 1)[2:]      # "26"

    # Intentar primero hoja de balance (datos en 1000t wse, mas completa)
    # luego hoja de produccion (datos en toneladas, solo produccion)
    sheet_candidates = _find_sheet_candidates(wb.sheetnames, yr_short, yr_next)

    for sheet_name in sheet_candidates:
        try:
            ws = wb[sheet_name]
            result = _extract_production_from_sheet(ws, campaign_year, sheet_name)
            if result is not None:
                logger.info("EC Sugar: %.3f Mt de hoja '%s'", result, sheet_name)
                wb.close()
                return result
        except Exception as e:
            logger.debug("EC XLSX hoja '%s': %s", sheet_name, e)

    wb.close()
    return None


def _find_sheet_candidates(sheetnames: list[str], yr_short: str, yr_next: str) -> list[str]:
    """Ordena hojas por relevancia: bilan > product, campaña actual primero.

    Para evitar falsos positivos entre hojas adyacentes (p.ej. "24-25 Apr 26"
    contiene "25" y "26" pero NO es la campaña 25/26), buscamos el string
    de campaña completo: "25-26" o "25/26".
    """
    priority = []
    secondary = []

    # El string de campaña que debe aparecer como bloque contiguo
    campaign_str = f"{yr_short}-{yr_next}"          # "25-26"
    campaign_str2 = f"{yr_short}/{yr_next}"          # "25/26"

    for name in sheetnames:
        name_low = name.lower()
        # Match exacto de campaña (no basta con que estén ambos años sueltos)
        is_current = (campaign_str in name_low or campaign_str2 in name_low)
        is_balance  = "bilan" in name_low or "balance" in name_low
        is_product  = "product" in name_low

        if is_current and (is_balance or is_product):
            priority.append(name)
        elif is_balance or is_product:
            secondary.append(name)

    # Si no hay candidatos con la campaña exacta, tomar todas las hojas de balance
    return priority + secondary


def _extract_production_from_sheet(ws, campaign_year: int, sheet_name: str) -> Optional[float]:
    """
    Busca la fila de produccion en una hoja y retorna el valor de la campaña.

    Estrategia A (hojas 'bilan'): valores en 1000 t wse, encontrar columna de
      la campaña actual y fila "Production".
    Estrategia B (hojas 'product'): valores en toneladas, buscar fila TOTAL.
    """
    sheet_low = sheet_name.lower()
    is_bilan   = "bilan" in sheet_low or "balance" in sheet_low
    is_product = "product" in sheet_low

    # Leer todas las filas en memoria (read_only necesita iteracion)
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)

    if not rows:
        return None

    if is_bilan:
        return _extract_from_bilan(rows, campaign_year)
    elif is_product:
        return _extract_from_product(rows)
    else:
        # Intentar ambas estrategias
        result = _extract_from_bilan(rows, campaign_year)
        if result is None:
            result = _extract_from_product(rows)
        return result


def _extract_from_bilan(rows: list, campaign_year: int) -> Optional[float]:
    """
    Hoja de balance: valores en 1000 t wse por campaña en columnas.
    Busca la columna de la campaña actual y la fila "Production".

    Estructura tipica:
      Fila 1: headers con nombres de campaña (p.ej. "17/18", "18/19" ... "25/26")
      Fila N: "Production" | val17 | val18 | ... | val_actual
    """
    yr_label_variants = [
        f"{str(campaign_year)[2:]}/{str(campaign_year+1)[2:]}",  # "25/26"
        f"{str(campaign_year)[2:]}-{str(campaign_year+1)[2:]}",  # "25-26"
        str(campaign_year),                                       # "2025"
    ]

    # Encontrar indice de columna de la campaña
    header_col_idx = None
    for row in rows[:10]:  # headers suelen estar en las primeras 10 filas
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).strip() if cell is not None else ""
            for lbl in yr_label_variants:
                if lbl in cell_str:
                    header_col_idx = col_idx
                    break
            if header_col_idx is not None:
                break
        if header_col_idx is not None:
            break

    # Si no encontramos header explicito, asumir ultima columna numerica
    if header_col_idx is None:
        # Tomar la ultima columna que tenga datos numericos en rango plausible
        header_col_idx = _guess_campaign_col(rows, campaign_year)

    if header_col_idx is None:
        return None

    # Buscar fila "Production"
    prod_keywords = {"production", "produccion", "production de sucre", "production totale"}
    for row in rows:
        label = str(row[0]).strip().lower() if row and row[0] is not None else ""
        if any(kw in label for kw in prod_keywords):
            try:
                val = row[header_col_idx]
                if val is None:
                    continue
                v = float(val)
                # Valores en 1000 t wse → rango esperado 12000–20000
                if 12_000 <= v <= 20_000:
                    return round(v / 1000, 3)
                # Algunos ficheros en t (12M–20M)
                if 12_000_000 <= v <= 20_000_000:
                    return round(v / 1_000_000, 3)
                # Algunos en Mt directamente
                if _EU_PROD_MIN <= v <= _EU_PROD_MAX:
                    return round(v, 3)
            except (ValueError, TypeError):
                pass

    return None


def _guess_campaign_col(rows: list, campaign_year: int) -> Optional[int]:
    """
    Heuristica: la columna de la campaña actual es la ultima con valores
    en rango plausible (12000–20000 kt) en la fila de produccion.
    """
    # Buscar fila que parezca produccion (primer valor numerico ~15000)
    for row in rows:
        if row and row[0] is not None:
            label = str(row[0]).lower()
            if "product" in label:
                # Iterar columnas de derecha a izquierda
                for col_idx in range(len(row) - 1, 0, -1):
                    try:
                        v = float(row[col_idx])
                        if 12_000 <= v <= 20_000:
                            return col_idx
                    except (ValueError, TypeError):
                        continue
    return None


def _extract_from_product(rows: list) -> Optional[float]:
    """
    Hoja de produccion por pais: valores en toneladas.
    Busca fila "TOTAL" y suma de todas las columnas numericas.
    """
    for row in rows:
        label = str(row[0]).strip().upper() if row and row[0] is not None else ""
        if label in ("TOTAL", "TOTAL UE", "TOTAL EU", "EU TOTAL"):
            # La produccion total esta en alguna columna numerica
            # Intentar la ultima o penultima columna con valor en rango (t)
            candidates = []
            for cell in row[1:]:
                try:
                    v = float(cell)
                    if 12_000_000 <= v <= 20_000_000:
                        candidates.append(v)
                except (ValueError, TypeError):
                    pass
            if candidates:
                return round(max(candidates) / 1_000_000, 3)

    return None


# ── Interfaz publica ───────────────────────────────────────────────────────────

def _current_campaign_year() -> int:
    """Campaña azucar UE: Oct 1 – Sep 30."""
    today = date.today()
    return today.year if today.month >= 10 else today.year - 1


def fetch_ec_sugar_production(
        campaign_year: Optional[int] = None,
        force_download: bool = False,
) -> Optional[dict]:
    """
    Obtiene produccion de azucar de la UE desde el XLSX de la CE.

    Retorna dict con {production_mt, campaign_year, data_date, source_url}
    o None si no hay datos.
    """
    yr = campaign_year or _current_campaign_year()

    xlsx = _get_xlsx_bytes(force_download=force_download)
    if xlsx is None:
        return None

    prod_mt = _parse_balance_sheet(xlsx, yr)
    if prod_mt is None:
        logger.warning("EC Sugar: XLSX descargado pero produccion no parseada para %d/%02d",
                       yr, (yr + 1) % 100)
        return None

    # Fecha del fichero en cache (si existe)
    data_date = date.today()
    if _CACHE_FILE.exists():
        data_date = date.fromtimestamp(_CACHE_FILE.stat().st_mtime)

    return {
        "production_mt": prod_mt,
        "campaign_year": yr,
        "data_date": data_date,
        "source_url": _AMO_SUGAR_PAGE,
    }


def eu_production_estimate(
        campaign_year: Optional[int] = None,
        usda_eu_mt: Optional[float] = None,
) -> tuple[Optional[float], str, float]:
    """
    Devuelve estimacion de produccion de azucar UE.

    Solo aplica override si:
      1. Dato EC disponible y fresco (<60 dias)
      2. Divergencia > 0.5 Mt vs USDA

    Returns:
        (production_mt, source_str, confidence)
        production_mt = None → usar USDA sin override
    """
    data = fetch_ec_sugar_production(campaign_year)

    if data is None:
        return None, "usda_eu_only", 0.80

    ec_mt = data["production_mt"]

    # Comprobar frescura del dato
    age_days = (date.today() - data["data_date"]).days
    if age_days > 60:
        logger.info("EC Sugar: dato de %d dias — demasiado viejo para override", age_days)
        return None, "ec_eu_stale", 0.75

    # Validar divergencia vs USDA
    if usda_eu_mt is not None:
        divergence = abs(ec_mt - usda_eu_mt)
        if divergence < 0.5:
            logger.info(
                "EU: EC=%.2f Mt vs USDA=%.2f Mt — divergencia <0.5 Mt, sin override",
                ec_mt, usda_eu_mt,
            )
            return None, "ec_confirms_usda", 0.85
        logger.info(
            "EU: EC=%.2f Mt vs USDA=%.2f Mt — divergencia %.2f Mt → override",
            ec_mt, usda_eu_mt, divergence,
        )

    return ec_mt, "ec_xlsx", 0.87
