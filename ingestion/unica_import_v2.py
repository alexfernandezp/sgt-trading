"""
Backfill UNICA desde el nuevo formato Excel (aa8c647f...).

Estructura nueva (diferente a unica_import.py v1):
  TB_MOAGEM      — cana molida (t neto por quincena)
  TB_PROD_ACUCAR — azucar producida (t neto)
  TB_PROD_EA     — etanol anidro (m3 neto)
  TB_PROD_EH     — etanol hidratado (m3 neto)

Columnas:
  [0] DATA_PUBLICACAO  (date = fecha quincena)
  [1] QUINZENA         (ej. '1a Abr')
  [2] SAFRA            (ej. '2025/2026')
  [3] COD_REGIAO       ('1'=Centro-Sul, '2'=SP, '3'=Demais)
  [4] COD_MP (MOAGEM) / COD_PRODUTO (PROD)
  [5] QUINZENAL (MOAGEM) / COD_MP (PROD)
  [6] QUINZENAL (PROD)

Uso:
    py scripts/ingest_unica_new_excel.py
"""
import logging
import os
from datetime import date

logger = logging.getLogger(__name__)

_ONEDRIVE = r"C:\Users\alejandro.fernandez\OneDrive - Sugar Global Trading\sgt_trading\data\Brazil Data"
_EXCEL_NEW = os.path.join(_ONEDRIVE, "aa8c647f13fbee79f297cc3ee62a716a.xlsx")


def _norm_safra(raw: str) -> str:
    """'2025/2026' -> '2025-2026'"""
    return raw.strip().replace("/", "-")


def _to_date(v) -> date | None:
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    if hasattr(v, "year"):
        return date(v.year, v.month, v.day)
    try:
        import datetime
        return datetime.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _read_moagem(wb) -> dict:
    """CS cana neta -> {(safra_norm, date): tons_float}"""
    out = {}
    for row in wb["TB_MOAGEM"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if str(row[3]).strip() != "1":   # solo CS
            continue
        if str(row[4]).strip() != "C":   # solo cana (C = Cana-de-acucar puro o combinado)
            continue
        val = row[5]
        if val is None:
            continue
        safra = _norm_safra(str(row[2]))
        d = _to_date(row[0])
        if d:
            out[(safra, d)] = float(val)
    return out


def _read_prod(wb, sheet: str, cod_prod: str) -> dict:
    """CS neto de un producto -> {(safra_norm, date): valor_float}"""
    out = {}
    for row in wb[sheet].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if str(row[3]).strip() != "1":        # solo CS
            continue
        if str(row[4]).strip() != cod_prod:   # filtrar por producto
            continue
        if str(row[5]).strip() != "C":        # solo cana (no milho)
            continue
        val = row[6]
        if val is None:
            continue
        safra = _norm_safra(str(row[2]))
        d = _to_date(row[0])
        if d:
            out[(safra, d)] = float(val)
    return out


def run_new_excel_backfill(session, excel_path: str = _EXCEL_NEW, commit: bool = True) -> dict:
    """
    Lee el nuevo Excel UNICA y hace upsert en unica_biweekly.
    Solo importa registros con cane_crushed_t no nulo.
    No sobreescribe atr_kg_ton ni sugar_mix_pct (no disponibles en este formato).
    """
    import openpyxl
    from sqlalchemy import text

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    moagem = _read_moagem(wb)
    acucar = _read_prod(wb, "TB_PROD_ACUCAR", "A")
    eth_a  = _read_prod(wb, "TB_PROD_EA",     "E_A")
    eth_h  = _read_prod(wb, "TB_PROD_EH",     "E_H")
    wb.close()

    all_keys = sorted(set(moagem.keys()) | set(acucar.keys()))

    upsert_sql = text("""
        INSERT INTO unica_biweekly
            (safra, quinzena_date, region,
             cane_crushed_t, sugar_t,
             ethanol_anidro_m3, ethanol_hidratado_m3, ethanol_total_m3)
        VALUES
            (:safra, :qdate, 'CS',
             :cane, :sugar,
             :eth_a, :eth_h, :eth_t)
        ON CONFLICT (safra, quinzena_date, region)
        DO UPDATE SET
            cane_crushed_t        = COALESCE(EXCLUDED.cane_crushed_t,        unica_biweekly.cane_crushed_t),
            sugar_t               = COALESCE(EXCLUDED.sugar_t,               unica_biweekly.sugar_t),
            ethanol_anidro_m3     = COALESCE(EXCLUDED.ethanol_anidro_m3,     unica_biweekly.ethanol_anidro_m3),
            ethanol_hidratado_m3  = COALESCE(EXCLUDED.ethanol_hidratado_m3,  unica_biweekly.ethanol_hidratado_m3),
            ethanol_total_m3      = COALESCE(EXCLUDED.ethanol_total_m3,      unica_biweekly.ethanol_total_m3)
    """)

    inserted = 0
    skipped  = 0
    for (safra, qdate) in all_keys:
        cane  = moagem.get((safra, qdate))
        sug   = acucar.get((safra, qdate))
        ea    = eth_a.get((safra, qdate))
        eh    = eth_h.get((safra, qdate))
        et    = (ea or 0) + (eh or 0) if (ea is not None or eh is not None) else None

        if cane is None and sug is None:
            skipped += 1
            continue

        try:
            session.execute(upsert_sql, {
                "safra": safra,
                "qdate": qdate,
                "cane":  int(round(cane))  if cane  is not None else None,
                "sugar": int(round(sug))   if sug   is not None else None,
                "eth_a": int(round(ea))    if ea    is not None else None,
                "eth_h": int(round(eh))    if eh    is not None else None,
                "eth_t": int(round(et))    if et    is not None else None,
            })
            inserted += 1
        except Exception as e:
            logger.warning("upsert error safra=%s date=%s: %s", safra, qdate, e)
            session.rollback()
            skipped += 1
            continue

    if commit:
        session.commit()

    return {"total": len(all_keys), "inserted": inserted, "skipped": skipped}


def _csv_date(dd_mm: str, start_year: int) -> date:
    """
    Convierte 'DD/MM' al date correcto dentro de una safra que empieza en start_year.
    La safra va de Apr-16 (start_year) hasta Apr-1 (start_year+1).
      16/04 → start_year-04-16  (seq=1)
      01/04 → (start_year+1)-04-01  (seq=24)
      Meses 1-3 → start_year+1
    """
    parts = dd_mm.strip().split("/")
    day, month = int(parts[0]), int(parts[1])
    if month < 4 or (month == 4 and day == 1):
        year = start_year + 1
    else:
        year = start_year
    return date(year, month, day)


def run_csv_mix_atr_backfill(
    session,
    safra: str = "2025-2026",
    mix_path: str | None = None,
    atr_path: str | None = None,
    commit: bool = True,
) -> dict:
    """
    Actualiza sugar_mix_pct, eth_mix_pct, atr_kg_ton para una safra ya importada.

    CSVs esperados (columnas: Quinzena, Safra Atual):
      MixAzucar2025-26.csv  — valores ej. '44.71%'
      ATR2025-26.csv        — valores ej. '103.46'

    eth_mix_pct = 100 - sugar_mix_pct  (resto va a etanol, como en UNICA)
    Solo hace UPDATE, no inserta filas nuevas.
    """
    import csv
    from sqlalchemy import text

    start_year = int(safra.split("-")[0])

    if mix_path is None:
        mix_path = os.path.join(_ONEDRIVE, "MixAzucar2025-26.csv")
    if atr_path is None:
        atr_path = os.path.join(_ONEDRIVE, "ATR2025-26.csv")

    # Leer mix
    mix_map: dict[date, float] = {}
    with open(mix_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            qz  = row.get("Quinzena", "").strip()
            val = row.get("Safra Atual", "").strip().rstrip("%")
            if not qz or not val:
                continue
            try:
                mix_map[_csv_date(qz, start_year)] = float(val)
            except Exception:
                logger.warning("mix: fila ignorada %s=%s", qz, val)

    # Leer ATR
    atr_map: dict[date, float] = {}
    with open(atr_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            qz  = row.get("Quinzena", "").strip()
            val = row.get("Safra Atual", "").strip()
            if not qz or not val:
                continue
            try:
                atr_map[_csv_date(qz, start_year)] = float(val)
            except Exception:
                logger.warning("atr: fila ignorada %s=%s", qz, val)

    update_sql = text("""
        UPDATE unica_biweekly
        SET
            sugar_mix_pct = :smix,
            eth_mix_pct   = :emix,
            atr_kg_ton    = :atr
        WHERE safra = :safra AND quinzena_date = :qdate AND region = 'CS'
    """)

    all_dates = sorted(set(mix_map) | set(atr_map))
    updated = 0
    for qdate in all_dates:
        smix = mix_map.get(qdate)
        emix = round(100.0 - smix, 2) if smix is not None else None
        atr  = atr_map.get(qdate)
        try:
            result = session.execute(update_sql, {
                "safra": safra,
                "qdate": qdate,
                "smix":  round(smix, 2) if smix is not None else None,
                "emix":  emix,
                "atr":   round(atr,  2) if atr  is not None else None,
            })
            if result.rowcount:
                updated += 1
        except Exception as e:
            logger.warning("update error date=%s: %s", qdate, e)

    if commit:
        session.commit()

    return {"dates": len(all_dates), "rows_updated": updated}


if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    sys.stdout.reconfigure(encoding="utf-8")
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    from database import SessionLocal
    with SessionLocal() as sess:
        stats = run_new_excel_backfill(sess)
    print("Backfill nuevo Excel UNICA:")
    print("  Total registros  :", stats["total"])
    print("  Upserted         :", stats["inserted"])
    print("  Skipped (null)   :", stats["skipped"])
    print()
    with SessionLocal() as sess:
        stats2 = run_csv_mix_atr_backfill(sess)
    print("Backfill Mix + ATR 2025-2026:")
    print("  Fechas procesadas:", stats2["dates"])
    print("  Filas actualizadas:", stats2["rows_updated"])
